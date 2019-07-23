#KonstTz
#rjshanahan@gmail.com
#23 July 2019

import openpyxl
import sys

class bcolors:
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def collect_content_for_search(sheetName):
    
    list_of_values= []
    
    wb_initialFile = openpyxl.load_workbook(sheetName, read_only=True, data_only=True)
    
    for sheet in wb_initialFile.worksheets:
        max_row=sheet.max_row
        max_column=sheet.max_column

        # iterate over all cells and rows
        for i in range(1,max_row+1):
            # iterate over all columns
            for j in range(1,max_column+1):
                # get particular cell value    
                cell_obj=sheet.cell(row=i,column=j)
                
                value = cell_obj.value
                if value != None :
                    list_of_values.append(value)
            
    return list_of_values

def serach_for_content (searching_for, searching_in):  
    not_found_list = []

    for value in searching_for:
        if value not in searching_in:
            not_found_list.append(value)

    return not_found_list

def print_data (not_found_list):

    for value in not_found_list:
        print (bcolors.WARNING, value, bcolors.FAIL,'- was not found',bcolors.ENDC)

if __name__ == "__main__":

    #load date from first argument
    searching_for = collect_content_for_search(sys.argv[1])

    #load date from seccond argument
    searching_in = collect_content_for_search(sys.argv[2])
    
    #get list of strings that wasn't found
    not_found_list = serach_for_content(searching_for, searching_in)

    if len(not_found_list) == 0 : 
        print(bcolors.OKGREEN,'Data was completelly checked. Every server was found',bcolors.ENDC)
    else:
        print_data(not_found_list)
